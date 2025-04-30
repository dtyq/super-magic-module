import csv
import os
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from app.filebase.parsers.base_parser import BaseParser
from app.filebase.vector.file_chunk import FileChunk
from app.logger import get_logger

logger = get_logger(__name__)


class ExcelParser(BaseParser):
    """
    Excel和CSV文件解析器，处理.xlsx、.xls和.csv文件
    """

    # 支持的文件扩展名
    SUPPORTED_EXTENSIONS: List[str] = ['.xlsx', '.xls', '.csv']

    def can_handle(self, file_path: str) -> bool:
        """
        检查是否可以处理该文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 是否可以处理
        """
        ext = os.path.splitext(file_path)[1].lower()
        return ext in self.SUPPORTED_EXTENSIONS

    def parse(self, file_path: str, metadata: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        解析Excel或CSV文件
        
        Args:
            file_path: 文件路径
            metadata: 附加元数据
            
        Returns:
            Dict[str, Any]: 解析结果，包括元数据、内容和FileChunk对象列表
        """
        if metadata is None:
            metadata = {}

        # 判断文件类型
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            return self._parse_csv(file_path, metadata)
        else:
            return self._parse_excel(file_path, metadata)

    def _handle_merged_cells(self, df: pd.DataFrame, file_path: str, sheet_name: str) -> Tuple[pd.DataFrame, List[Tuple]]:
        """
        处理Excel文件中的合并单元格，填充合并单元格中的空值

        Args:
            df: pandas DataFrame
            file_path: Excel文件路径
            sheet_name: 工作表名称

        Returns:
            Tuple: 处理后的DataFrame和合并单元格范围列表
        """
        try:
            # 对于非xlsx文件，不使用openpyxl处理
            if not file_path.lower().endswith('.xlsx'):
                return df, []

            # 使用openpyxl读取合并单元格信息
            wb = load_workbook(file_path, data_only=True)

            if sheet_name not in wb.sheetnames:
                logger.warning(f"工作表 '{sheet_name}' 在openpyxl中不存在")
                return df, []

            ws = wb[sheet_name]

            # 提取所有合并单元格的范围
            merged_ranges = []
            for merged_cell_range in ws.merged_cells.ranges:
                merged_ranges.append(merged_cell_range.bounds)

                # 获取合并区域的首个单元格值
                start_cell = merged_cell_range.start_cell
                start_value = start_cell.value

                # 如果首个单元格没有值，则不需要填充
                if start_value is None:
                    continue

                # 获取合并区域的边界(bounds返回1-based索引)
                min_row, min_col, max_row, max_col = merged_cell_range.bounds

                # 填充合并区域中的所有单元格
                # 注意：pandas的索引是0-based，而openpyxl的索引是1-based，需要调整
                for row in range(min_row - 1, max_row):
                    for col in range(min_col - 1, max_col):
                        # 跳过起始单元格(已经有值)
                        if row == min_row - 1 and col == min_col - 1:
                            continue
                        # 使用iat进行快速赋值(整数位置索引)
                        try:
                            df.iat[row, col] = start_value
                        except IndexError:
                            # 处理可能的索引越界情况
                            pass

            return df, merged_ranges

        except Exception as e:
            logger.error(f"处理合并单元格时出错: {e!s}")
            # 如果openpyxl处理失败，回退到简单填充方法
            df_filled = df.fillna(method='ffill', axis=0)
            return df_filled, []

    def _parse_excel(self, file_path: str, metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        解析Excel文件
        
        Args:
            file_path: 文件路径
            metadata: 附加元数据
            
        Returns:
            Dict[str, Any]: 解析结果，包括元数据、内容和FileChunk对象列表
        """
        # 添加解析器信息到元数据
        metadata['parser'] = 'excel'
        metadata['file_type'] = 'excel'
        metadata['has_merged_cells'] = False
        merged_cells_info = {}

        try:
            # 使用pandas读取Excel文件
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            metadata['sheet_count'] = len(sheet_names)
            metadata['sheet_names'] = sheet_names

            all_content = []
            file_chunks = []

            # 根据字符长度决定如何分块，确保不超过嵌入模型的上下文限制
            max_chunk_chars = 4000  # 每个块包含的最大字符数，根据模型限制调整
            max_single_row_chars = 8000  # 单行最大字符数，如果超过此值将记录警告

            # 处理每个工作表
            total_chunk_index = 0

            for sheet_index, sheet_name in enumerate(sheet_names):
                # 使用pandas读取Excel工作表
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                # 处理合并单元格
                df, merged_ranges = self._handle_merged_cells(df, file_path, sheet_name)

                # 记录合并单元格信息
                if merged_ranges:
                    metadata['has_merged_cells'] = True
                    merged_cells_info[sheet_name] = merged_ranges

                # 获取行列数
                rows, cols = df.shape

                # 为当前工作表添加元数据
                sheet_metadata = {
                    'sheet_name': sheet_name,
                    'sheet_index': sheet_index,
                    'rows': rows,
                    'columns': cols,
                    'content_type': 'excel_sheet',
                    'merged_cells_count': len(merged_ranges) if merged_ranges else 0
                }

                # 工作表标题
                sheet_title = f"--- 工作表: {sheet_name} ({rows}行 x {cols}列) ---"

                # 将DataFrame转换为列表格式，便于统一格式化处理
                data_rows = []
                # 添加表头
                header_row = [sheet_title]  # 工作表标题作为第一行
                data_rows.append(header_row)

                # 添加列名行
                column_names = df.columns.tolist()
                data_rows.append(column_names)

                # 添加数据行
                for idx, row in df.iterrows():
                    data_rows.append([idx] + row.tolist())  # 包含行索引

                # 统一格式化为竖线分隔的字符串
                formatted_rows = []
                for row in data_rows:
                    # 确保所有值都是字符串
                    str_row = [str(val) for val in row]
                    formatted_rows.append(" | ".join(str_row))

                # 转换为字符串表示
                sheet_content = "\n".join(formatted_rows)
                all_content.append(sheet_content)

                # 检查每行是否有超出单行最大长度的情况
                for i, row in enumerate(formatted_rows):
                    if len(row) > max_single_row_chars:
                        logger.warning(f"Row {i} in sheet '{sheet_name}' of Excel file {file_path} exceeds max single row length ({len(row)} > {max_single_row_chars} chars)")

                # 标题行和表头行
                title_row = formatted_rows[0]
                header_row = formatted_rows[1]

                if len(sheet_content) <= max_chunk_chars:
                    # 如果整个工作表内容较小，作为一个块
                    file_chunk = FileChunk(
                        text=sheet_content,
                        file_metadata=metadata,
                        chunk_metadata={
                            **sheet_metadata,
                            'row_range': f"1-{rows}",
                            'is_complete_sheet': True
                        },
                        chunk_index=total_chunk_index,
                        total_chunks=-1  # 暂时未知总块数
                    )
                    file_chunks.append(file_chunk)
                    total_chunk_index += 1
                else:
                    # 如果工作表内容较大，按内容长度分块，保持行的完整性
                    current_chunk = [title_row, header_row]  # 每个块都从标题和表头开始
                    current_length = len(title_row) + len(header_row) + 1  # +1 是换行符
                    current_start_row = 1  # 数据从第1行开始（表头是第1行，标题是第0行）
                    sheet_chunk_index = 0

                    # 从第2行开始（跳过标题和表头）
                    for i, row in enumerate(formatted_rows[2:], 2):
                        row_length = len(row) + 1  # +1 是换行符

                        # 如果当前块已经有内容，且添加当前行会超出长度限制，先保存当前块
                        if len(current_chunk) > 2 and current_length + row_length > max_chunk_chars:
                            # 合并为文本
                            chunk_content = "\n".join(current_chunk)

                            file_chunk = FileChunk(
                                text=chunk_content,
                                file_metadata=metadata,
                                chunk_metadata={
                                    **sheet_metadata,
                                    'row_range': f"{current_start_row}-{i-1}",
                                    'is_complete_sheet': False,
                                    'sheet_chunk_index': sheet_chunk_index
                                },
                                chunk_index=total_chunk_index,
                                total_chunks=-1  # 暂时未知总块数
                            )
                            file_chunks.append(file_chunk)
                            total_chunk_index += 1

                            # 重置为新块，保留标题和表头
                            current_chunk = [title_row, header_row]
                            current_length = len(title_row) + len(header_row) + 1
                            current_start_row = i
                            sheet_chunk_index += 1

                        # 添加当前行到块（无论长度如何，保持行的完整性）
                        current_chunk.append(row)
                        current_length += row_length

                        # 如果单行非常长，记录警告但保持完整性
                        if row_length > max_chunk_chars:
                            logger.warning(f"Row {i} in sheet '{sheet_name}' of Excel file {file_path} is longer than max_chunk_chars ({row_length} > {max_chunk_chars}), keeping it intact")

                    # 处理最后一个块
                    if len(current_chunk) > 2:  # 确保不只有标题和表头
                        chunk_content = "\n".join(current_chunk)

                        file_chunk = FileChunk(
                            text=chunk_content,
                            file_metadata=metadata,
                            chunk_metadata={
                                **sheet_metadata,
                                'row_range': f"{current_start_row}-{rows}",
                                'is_complete_sheet': False,
                                'sheet_chunk_index': sheet_chunk_index
                            },
                            chunk_index=total_chunk_index,
                            total_chunks=-1  # 暂时未知总块数
                        )
                        file_chunks.append(file_chunk)
                        total_chunk_index += 1

            # 合并所有工作表内容
            content = "\n\n".join(all_content)

            # 更新所有块的total_chunks
            total_chunks = len(file_chunks)
            for chunk in file_chunks:
                chunk.total_chunks = total_chunks

            # 添加合并单元格信息到元数据
            if metadata['has_merged_cells']:
                metadata['merged_cells_info'] = merged_cells_info

            return {
                'metadata': metadata,
                'content': content,
                'chunks': file_chunks
            }

        except Exception as e:
            logger.error(f"解析Excel文件 {file_path} 时出错: {e!s}")
            return {
                'metadata': metadata,
                'content': f"Excel文件 {file_path} 解析失败: {e!s}",
                'chunks': []
            }

    def _parse_csv(self, file_path: str, metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        解析CSV文件
        
        Args:
            file_path: 文件路径
            metadata: 附加元数据
            
        Returns:
            Dict[str, Any]: 解析结果，包括元数据、内容和FileChunk对象列表
        """
        # 添加解析器信息到元数据
        metadata['parser'] = 'csv'
        metadata['file_type'] = 'csv'

        # 读取CSV文件
        try:
            # 尝试检测编码和分隔符
            encoding = self._detect_encoding(file_path)
            delimiter = self._detect_delimiter(file_path, encoding)

            # 使用pandas读取CSV文件
            df = pd.read_csv(file_path, encoding=encoding, delimiter=delimiter, on_bad_lines='skip')

            # 获取行列数添加到元数据
            rows, cols = df.shape
            metadata['rows'] = rows
            metadata['columns'] = cols
            metadata['column_names'] = df.columns.tolist()
            metadata['encoding'] = encoding
            metadata['delimiter'] = delimiter

            # 将DataFrame转换为列表格式，便于统一格式化处理
            data_rows = []
            # 添加表头
            data_rows.append(df.columns.tolist())
            # 添加数据行
            for _, row in df.iterrows():
                data_rows.append(row.tolist())

            # 统一格式化为竖线分隔的字符串
            formatted_rows = []
            for row in data_rows:
                # 确保所有值都是字符串
                str_row = [str(val) for val in row]
                formatted_rows.append(" | ".join(str_row))

            # 转换为字符串表示
            content = "\n".join(formatted_rows)

            # 创建 FileChunk 对象列表
            file_chunks = []

            # 根据字符长度决定如何分块，确保不超过嵌入模型的上下文限制
            max_chunk_chars = 4000  # 每个块包含的最大字符数，根据模型限制调整
            max_single_row_chars = 8000  # 单行最大字符数，如果超过此值将记录警告

            # 确保分块，即使数据量很小也进行分块
            # 表头行
            header_row = formatted_rows[0]

            # 检查每行是否有超出单行最大长度的情况
            for i, row in enumerate(formatted_rows):
                if len(row) > max_single_row_chars:
                    logger.warning(f"Row {i} in CSV file {file_path} exceeds max single row length ({len(row)} > {max_single_row_chars} chars)")

            if len(content) <= max_chunk_chars:
                # 如果整个内容较小，整个文件作为一个块
                file_chunk = FileChunk(
                    text=content,
                    file_metadata=metadata,
                    chunk_metadata={
                        'row_range': f"1-{rows}",
                        'is_complete_file': True
                    },
                    chunk_index=0,
                    total_chunks=1
                )
                file_chunks.append(file_chunk)
            else:
                # 如果内容较大，按内容长度分块，保持行的完整性
                current_chunk = [header_row]  # 每个块都从表头开始
                current_length = len(header_row)
                current_start_row = 1  # 数据从第1行开始（表头是第0行）
                chunk_index = 0

                # 从第1行开始（跳过表头）
                for i, row in enumerate(formatted_rows[1:], 1):
                    row_length = len(row) + 1  # +1 是换行符

                    # 如果当前块已经有内容，且添加当前行会超出长度限制，先保存当前块
                    if len(current_chunk) > 1 and current_length + row_length > max_chunk_chars:
                        # 合并为文本
                        chunk_content = "\n".join(current_chunk)

                        file_chunk = FileChunk(
                            text=chunk_content,
                            file_metadata=metadata,
                            chunk_metadata={
                                'row_range': f"{current_start_row}-{i-1}",
                                'is_complete_file': False
                            },
                            chunk_index=chunk_index,
                            total_chunks=-1  # 暂时未知总块数
                        )
                        file_chunks.append(file_chunk)

                        # 重置为新块，保留表头
                        current_chunk = [header_row]
                        current_length = len(header_row)
                        current_start_row = i
                        chunk_index += 1

                    # 添加当前行到块（无论长度如何，保持行的完整性）
                    current_chunk.append(row)
                    current_length += row_length

                    # 如果单行非常长，记录警告但保持完整性
                    if row_length > max_chunk_chars:
                        logger.warning(f"Row {i} in CSV file {file_path} is longer than max_chunk_chars ({row_length} > {max_chunk_chars}), keeping it intact")

                # 处理最后一个块
                if len(current_chunk) > 1:  # 确保不只有表头
                    chunk_content = "\n".join(current_chunk)

                    file_chunk = FileChunk(
                        text=chunk_content,
                        file_metadata=metadata,
                        chunk_metadata={
                            'row_range': f"{current_start_row}-{rows}",
                            'is_complete_file': False
                        },
                        chunk_index=chunk_index,
                        total_chunks=-1  # 暂时未知总块数
                    )
                    file_chunks.append(file_chunk)

                # 更新所有块的total_chunks
                total_chunks = len(file_chunks)
                for chunk in file_chunks:
                    chunk.total_chunks = total_chunks

            return {
                'metadata': metadata,
                'content': content,
                'chunks': file_chunks
            }
        except Exception as e:
            logger.error(f"Error parsing CSV file {file_path}: {e!s}")

            # 如果pandas方法失败，尝试使用标准csv库
            try:
                with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
                    csv_reader = csv.reader(file)
                    rows = list(csv_reader)

                    if rows:
                        # 获取行列数添加到元数据
                        metadata['rows'] = len(rows)
                        metadata['columns'] = len(rows[0]) if rows else 0

                        # 格式化内容
                        content = []
                        for row in rows:
                            content.append(" | ".join(row))

                        content_str = "\n".join(content)

                        # 创建FileChunk对象
                        file_chunks = []

                        # 根据字符长度决定如何分块
                        max_chunk_chars = 4000  # 每个块包含的最大字符数
                        max_single_row_chars = 8000  # 单行最大字符数，如果超过此值将记录警告

                        # 表头行
                        header_row = content[0] if content else ""

                        # 检查每行是否有超出单行最大长度的情况
                        for i, row in enumerate(content):
                            if len(row) > max_single_row_chars:
                                logger.warning(f"Row {i} in CSV file {file_path} exceeds max single row length ({len(row)} > {max_single_row_chars} chars)")

                        if len(content_str) <= max_chunk_chars:
                            # 如果整个内容较小，整个文件作为一个块
                            file_chunk = FileChunk(
                                text=content_str,
                                file_metadata=metadata,
                                chunk_metadata={
                                    'row_range': f"1-{len(rows)}",
                                    'is_complete_file': True,
                                    'fallback_method': True
                                },
                                chunk_index=0,
                                total_chunks=1
                            )
                            file_chunks.append(file_chunk)
                        else:
                            # 如果内容较大，按内容长度分块，保持行的完整性
                            current_chunk = [header_row]  # 每个块都从表头开始
                            current_length = len(header_row)
                            current_start_row = 1  # 数据从第1行开始（表头是第0行）
                            chunk_index = 0

                            # 从第1行开始（跳过表头）
                            for i, row in enumerate(content[1:], 1):
                                row_length = len(row) + 1  # +1 是换行符

                                # 如果当前块已经有内容，且添加当前行会超出长度限制，先保存当前块
                                if len(current_chunk) > 1 and current_length + row_length > max_chunk_chars:
                                    # 合并为文本
                                    chunk_content = "\n".join(current_chunk)

                                    file_chunk = FileChunk(
                                        text=chunk_content,
                                        file_metadata=metadata,
                                        chunk_metadata={
                                            'row_range': f"{current_start_row}-{i-1}",
                                            'is_complete_file': False,
                                            'fallback_method': True
                                        },
                                        chunk_index=chunk_index,
                                        total_chunks=-1  # 暂时未知总块数
                                    )
                                    file_chunks.append(file_chunk)

                                    # 重置为新块，保留表头
                                    current_chunk = [header_row]
                                    current_length = len(header_row)
                                    current_start_row = i
                                    chunk_index += 1

                                # 添加当前行到块（无论长度如何，保持行的完整性）
                                current_chunk.append(row)
                                current_length += row_length

                                # 如果单行非常长，记录警告但保持完整性
                                if row_length > max_chunk_chars:
                                    logger.warning(f"Row {i} in CSV file {file_path} is longer than max_chunk_chars ({row_length} > {max_chunk_chars}), keeping it intact")

                            # 处理最后一个块
                            if len(current_chunk) > 1:  # 确保不只有表头
                                chunk_content = "\n".join(current_chunk)

                                file_chunk = FileChunk(
                                    text=chunk_content,
                                    file_metadata=metadata,
                                    chunk_metadata={
                                        'row_range': f"{current_start_row}-{len(rows)-1}",
                                        'is_complete_file': False,
                                        'fallback_method': True
                                    },
                                    chunk_index=chunk_index,
                                    total_chunks=-1  # 暂时未知总块数
                                )
                                file_chunks.append(file_chunk)

                            # 更新所有块的total_chunks
                            total_chunks = len(file_chunks)
                            for chunk in file_chunks:
                                chunk.total_chunks = total_chunks

                        return {
                            'metadata': metadata,
                            'content': content_str,
                            'chunks': file_chunks
                        }
            except Exception as e2:
                logger.error(f"Fallback CSV parsing also failed for {file_path}: {e2!s}")

            # 返回空内容
            return {
                'metadata': metadata,
                'content': "",
                'chunks': []
            }

    def _detect_encoding(self, file_path: str) -> str:
        """
        尝试检测CSV文件的编码
        
        Args:
            file_path: 文件路径
            
        Returns:
            str: 编码
        """
        encodings = ['utf-8', 'gbk', 'latin1', 'cp1252']

        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as file:
                    file.read(1024)  # 读取一小部分内容测试
                return encoding
            except UnicodeDecodeError:
                continue

        # 默认返回utf-8并忽略错误
        return 'utf-8'

    def _detect_delimiter(self, file_path: str, encoding: str) -> str:
        """
        尝试检测CSV文件的分隔符
        
        Args:
            file_path: 文件路径
            encoding: 文件编码
            
        Returns:
            str: 分隔符
        """
        delimiters = [',', ';', '\t', '|']

        try:
            with open(file_path, 'r', encoding=encoding, errors='replace') as file:
                sample = file.readline()

                # 对每个可能的分隔符计算出现次数
                counts = {delimiter: sample.count(delimiter) for delimiter in delimiters}

                # 选择出现次数最多的分隔符
                max_delimiter = max(counts.items(), key=lambda x: x[1])

                # 如果分隔符至少出现过一次，使用它
                if max_delimiter[1] > 0:
                    return max_delimiter[0]
        except:
            pass

        # 默认使用逗号
        return ',' 
