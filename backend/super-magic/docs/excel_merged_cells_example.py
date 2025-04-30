#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel合并单元格处理示例

此示例展示了如何处理Excel文件中的合并单元格，以确保数据被正确读取和处理。
通过结合pandas和openpyxl的功能，实现对合并单元格的准确识别和内容填充。
"""

import os
import pandas as pd
from openpyxl import load_workbook
from typing import Dict, Any, Optional, List, Tuple


def fill_merged_cells_simple(df: pd.DataFrame) -> pd.DataFrame:
    """
    使用pandas自带的fillna方法填充合并单元格中的NaN值
    
    Args:
        df: 包含合并单元格数据的DataFrame
        
    Returns:
        处理后的DataFrame，合并单元格已被填充
    """
    # 按行向下填充（处理垂直合并单元格）
    df_filled = df.fillna(method='ffill', axis=0)
    
    # 可选：按列向右填充（处理水平合并单元格）
    # df_filled = df_filled.fillna(method='ffill', axis=1)
    
    return df_filled


def get_merged_cell_ranges(file_path: str, sheet_name: str) -> List[Tuple]:
    """
    获取Excel文件中特定工作表的所有合并单元格范围
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        
    Returns:
        合并单元格范围列表，每个范围为(min_row, min_col, max_row, max_col)
    """
    # 加载工作簿，不建议使用read_only=True，会导致无法访问合并单元格信息
    wb = load_workbook(file_path, data_only=True)
    
    # 获取指定工作表
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表 '{sheet_name}' 不存在")
        
    ws = wb[sheet_name]
    
    # 提取所有合并单元格的范围
    merged_ranges = []
    for merged_cell_range in ws.merged_cells.ranges:
        merged_ranges.append(merged_cell_range.bounds)
    
    return merged_ranges


def fill_merged_cells_accurate(file_path: str, sheet_name: str = None) -> pd.DataFrame:
    """
    精确处理Excel文件中的合并单元格，使用openpyxl获取合并单元格信息
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，默认为None(第一个工作表)
        
    Returns:
        处理后的DataFrame，合并单元格已被准确填充
    """
    # 首先使用pandas读取Excel文件
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # 如果未指定sheet_name，获取第一个工作表名称
    if sheet_name is None:
        excel_file = pd.ExcelFile(file_path)
        sheet_name = excel_file.sheet_names[0]
    
    try:
        # 使用openpyxl获取合并单元格信息
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        
        # 处理每个合并单元格区域
        for merged_range in ws.merged_cells.ranges:
            # 获取合并区域的首个单元格值(合并区域的单元格值都存储在首个单元格中)
            start_cell = merged_range.start_cell
            start_value = start_cell.value
            
            # 如果首个单元格没有值，则不需要填充
            if start_value is None:
                continue
                
            # 获取合并区域的边界(bounds返回1-based索引)
            min_row, min_col, max_row, max_col = merged_range.bounds
            
            # 填充合并区域中的所有单元格
            # 注意：pandas的索引是0-based，而openpyxl的索引是1-based，需要调整
            for row in range(min_row - 1, max_row):
                for col in range(min_col - 1, max_col):
                    # 跳过起始单元格(已经有值)
                    if row == min_row - 1 and col == min_col - 1:
                        continue
                    # 使用iat进行快速赋值(整数位置索引)
                    try:
                        df.iat[row, col] = start_value
                    except IndexError:
                        # 处理可能的索引越界情况
                        pass
        
        return df
    
    except Exception as e:
        print(f"处理合并单元格时出错: {str(e)}")
        # 如果openpyxl处理失败，回退到简单填充方法
        return fill_merged_cells_simple(df)


def enhance_excel_parser_for_merged_cells(file_path: str, sheet_name: str = None) -> Dict[str, Any]:
    """
    增强版Excel解析器，用于filebase系统，处理合并单元格问题
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，默认为None(处理所有工作表)
        
    Returns:
        Dict: 包含解析后数据和元数据的字典
    """
    excel_file = pd.ExcelFile(file_path)
    sheet_names = excel_file.sheet_names if sheet_name is None else [sheet_name]
    
    all_content = []
    file_chunks = []
    metadata = {
        'parser': 'excel_enhanced',
        'file_type': 'excel',
        'sheet_count': len(sheet_names),
        'sheet_names': sheet_names,
        'has_merged_cells': False  # 默认无合并单元格
    }
    
    merged_cells_info = {}
    
    for sheet_index, current_sheet_name in enumerate(sheet_names):
        try:
            # 使用增强的方法处理合并单元格
            df = fill_merged_cells_accurate(file_path, current_sheet_name)
            
            # 获取合并单元格范围信息
            merged_ranges = get_merged_cell_ranges(file_path, current_sheet_name)
            
            if merged_ranges:
                metadata['has_merged_cells'] = True
                merged_cells_info[current_sheet_name] = merged_ranges
            
            # 获取行列数
            rows, cols = df.shape
            
            # 为当前工作表添加元数据
            sheet_metadata = {
                'sheet_name': current_sheet_name,
                'sheet_index': sheet_index,
                'rows': rows,
                'columns': cols,
                'content_type': 'excel_sheet',
                'merged_cells_count': len(merged_ranges)
            }
            
            # 将DataFrame转换为文本内容并添加到所有内容中
            sheet_content = format_df_to_text(df, sheet_metadata)
            all_content.append(sheet_content)
            
            # 在这里可以添加文件分块逻辑，类似filebase现有的实现
            # ...
            
        except Exception as e:
            print(f"处理工作表 {current_sheet_name} 时出错: {str(e)}")
    
    # 如果有合并单元格，添加到元数据
    if metadata['has_merged_cells']:
        metadata['merged_cells_info'] = merged_cells_info
    
    # 合并所有工作表内容
    content = "\n\n".join(all_content)
    
    return {
        'metadata': metadata,
        'content': content,
        'chunks': file_chunks  # 实际实现中需要填充此列表
    }


def format_df_to_text(df: pd.DataFrame, metadata: Dict[str, Any]) -> str:
    """
    将DataFrame格式化为文本以便于显示和处理
    
    Args:
        df: 数据DataFrame
        metadata: 工作表元数据
        
    Returns:
        格式化后的文本内容
    """
    sheet_title = f"--- 工作表: {metadata['sheet_name']} ({metadata['rows']}行 x {metadata['columns']}列) ---"
    
    # 转换DataFrame为字符串表示
    content_lines = [sheet_title]
    
    # 添加列名
    column_names = " | ".join([str(col) for col in df.columns])
    content_lines.append(column_names)
    
    # 添加数据行
    for idx, row in df.iterrows():
        row_str = " | ".join([str(val) for val in row])
        content_lines.append(f"{idx} | {row_str}")
    
    return "\n".join(content_lines)


if __name__ == "__main__":
    # 使用示例
    sample_file = "example.xlsx"
    
    if os.path.exists(sample_file):
        print("示例1: 使用简单填充方法")
        df1 = pd.read_excel(sample_file)
        filled_df1 = fill_merged_cells_simple(df1)
        print(filled_df1.head())
        
        print("\n示例2: 使用精确填充方法")
        filled_df2 = fill_merged_cells_accurate(sample_file)
        print(filled_df2.head())
        
        print("\n示例3: 使用增强版Excel解析器")
        result = enhance_excel_parser_for_merged_cells(sample_file)
        print(f"元数据: {result['metadata']}")
        print(f"是否包含合并单元格: {result['metadata']['has_merged_cells']}")
    else:
        print(f"示例文件 {sample_file} 不存在，请创建后再运行此脚本。")
        print("示例文件应包含一些合并单元格以展示功能。") 